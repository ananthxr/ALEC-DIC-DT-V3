
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Load the updated CSV file
file_path = 'Y:\ALEC DIC DT V3\Assets\Sensor Excels\RoomData -test Gemini.csv'
df = pd.read_csv(file_path)

# Create a new Excel workbook
wb = Workbook()
ws = wb.active

# Append the header
ws.append(df.columns.tolist())

# Define the green fill
green_fill = PatternFill(start_color='00FF00',
                           end_color='00FF00',
                           fill_type='solid')

# Append the data and color the rows
for index, row in df.iterrows():
    # openpyxl needs a list of values
    ws.append(row.astype(str).tolist())
    # Check if the row was modified (Type is 'room' and Building is not empty)
    if row['Type'] == 'room' and isinstance(row['Building'], str) and row['Building'] != '':
        for cell in ws[ws.max_row]:
            cell.fill = green_fill

# Save the new Excel file
output_path = 'Y:\ALEC DIC DT V3\Assets\Sensor Excels\RoomData -test Gemini_colored.xlsx'
wb.save(output_path)

print(f"Colored Excel file saved at '{output_path}'")
